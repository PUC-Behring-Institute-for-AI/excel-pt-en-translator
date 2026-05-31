#!/usr/bin/env python3
"""
Translate Excel file from Portuguese to English.

Supports two providers via --provider:

  opencode-zen  (default) — OpenCode Zen gateway, OpenAI-compatible.
                            Free model: opencode/big-pickle
                            Key: OPENCODE_ZEN_API_KEY env var or --api-key

  gemini                  — Google Gemini API.
                            Tested model: gemini-2.5-flash
                            Key: GEMINI_API_KEY env var or --api-key

Usage:
    python translate_excel.py input.xlsx output.xlsx                          # opencode-zen default
    python translate_excel.py input.xlsx output.xlsx --provider gemini --api-key AIza...
    python translate_excel.py input.xlsx output.xlsx --list-models
"""

import argparse
import json
import os
import re
import sys
import time
from datetime import datetime

import openpyxl
from openpyxl.cell.cell import MergedCell

# Provider constants
PROVIDER_OPENCODE = "opencode-zen"
PROVIDER_GEMINI   = "gemini"

OPENCODE_BASE_URL   = "https://opencode.ai/zen/v1"
OPENCODE_DEFAULT    = "opencode/big-pickle"
GEMINI_DEFAULT      = "gemini-2.5-flash"


# ─── Progress logger ────────────────────────────────────────────────────────

class Progress:
    def __init__(self, total_cells: int, total_batches: int):
        self.total_cells   = total_cells
        self.total_batches = total_batches
        self.cells_done    = 0
        self.batches_done  = 0
        self.start_time    = time.time()

    def _elapsed(self) -> str:
        s = int(time.time() - self.start_time)
        return f"{s // 60:02d}:{s % 60:02d}"

    def _eta(self) -> str:
        elapsed = time.time() - self.start_time
        if self.cells_done == 0:
            return "--:--"
        rate = self.cells_done / elapsed
        remaining = self.total_cells - self.cells_done
        eta_s = int(remaining / rate) if rate > 0 else 0
        return f"{eta_s // 60:02d}:{eta_s % 60:02d}"

    def start_batch(self, batch_num: int, sheet: str):
        pct = self.cells_done / self.total_cells * 100 if self.total_cells else 0
        bar_len = 25
        filled = int(bar_len * pct / 100)
        bar = "█" * filled + "░" * (bar_len - filled)
        print(
            f"\r[{bar}] {pct:5.1f}%  "
            f"batch {batch_num}/{self.total_batches}  "
            f"cells {self.cells_done}/{self.total_cells}  "
            f"elapsed {self._elapsed()}  ETA {self._eta()}  "
            f"sheet: {sheet[:28]}",
            end="  ",
            flush=True,
        )

    def finish_batch(self, batch_size: int):
        self.cells_done += batch_size
        self.batches_done += 1

    def done(self):
        elapsed = int(time.time() - self.start_time)
        print(
            f"\r[{'█' * 25}] 100.0%  "
            f"batch {self.total_batches}/{self.total_batches}  "
            f"cells {self.total_cells}/{self.total_cells}  "
            f"elapsed {elapsed // 60:02d}:{elapsed % 60:02d}  ETA 00:00"
            + " " * 40
        )


# ─── Shared helpers ──────────────────────────────────────────────────────────

def should_translate(cell) -> bool:
    if isinstance(cell, MergedCell):
        return False
    value = cell.value
    if value is None or not isinstance(value, str):
        return False
    stripped = value.strip()
    if not stripped:
        return False
    if stripped.startswith(("http://", "https://")):
        return False
    return True


def _build_prompt(texts: list) -> str:
    texts_json = json.dumps(texts, ensure_ascii=False, indent=2)
    return f"""You are a professional translator. Translate the following texts from Portuguese to English.

STRICT RULES:
1. Proper nouns (person names, place names) → keep unchanged.
2. Institution / brand names → keep unchanged.
   Examples: PUC-Rio, ICA, CCEC, ECOA, BI MASTER, ExperIA, AcademIA, AI MASTER,
             AI LAB, Lattes, MCTI, IRI, Power BI, AutoGen, CrewAI, LlamaIndex,
             LangChain, HuggingFace, Google Analytics, Figma, CapCut, Zoom, Adobe.
3. Course codes → keep unchanged (e.g. BIA1001, MEC2007, CIS2114, CTN1408).
4. URLs → keep exactly unchanged.
5. Standard English / institution acronyms → keep unchanged.
   Examples: AI, ML, LLM, RAG, NLP, MoE, MCP, XAI, IDP, API, BI, ChatGPT, GPT.
6. English technical terms already in common use → keep in English.
   Examples: Machine Learning, Deep Learning, Bootcamp, Big Data, Dashboard,
             Streaming, Prompt Engineering, Embeddings, Pipeline, Chatbot.
7. Preserve internal line breaks (\\n) within each segment.
8. Return ONLY a valid JSON array with EXACTLY {len(texts)} strings, same order.
   No markdown, no commentary outside the JSON.

Input ({len(texts)} texts):
{texts_json}

Output (JSON array, {len(texts)} strings):"""


def _parse_response(raw: str, expected: int, texts: list) -> list:
    """Parse a JSON array from model output. Falls back to originals on failure."""
    if raw.startswith("```"):
        lines = raw.splitlines()
        raw = "\n".join(lines[1:-1]).strip()
    try:
        result = json.loads(raw)
        if not isinstance(result, list) or len(result) != expected:
            raise ValueError(f"Got {len(result) if isinstance(result, list) else '?'}, expected {expected}")
        return [str(t) for t in result]
    except (json.JSONDecodeError, ValueError) as exc:
        print(f"\n  WARNING: could not parse response ({exc}) — keeping originals for this batch.")
        return texts


def _handle_error(exc: Exception, attempt: int, max_retries: int) -> float | None:
    """
    Decide what to do with an exception.
    Returns a sleep duration if we should retry, None if we should use originals,
    or raises SystemExit for fatal errors.
    """
    err_str = str(exc)

    if "429" in err_str or "rate" in err_str.lower() or "RESOURCE_EXHAUSTED" in err_str:
        if attempt < max_retries - 1:
            delay = 62.0
            match = re.search(r"retry[^\d]*(\d+(?:\.\d+)?)\s*s", err_str, re.I)
            if match:
                delay = float(match.group(1)) + 3
            print(f"\n  [rate limit] waiting {delay:.0f}s (attempt {attempt + 1}/{max_retries})…", flush=True)
            return delay
        print(f"\n  WARNING: rate limit persisted after {max_retries} retries — keeping originals.")
        return None

    if any(code in err_str for code in ("401", "403", "404", "NOT_FOUND", "PERMISSION_DENIED", "Unauthorized")):
        print(f"\n\nFATAL: {exc}\n")
        sys.exit(1)

    print(f"\n  WARNING: unexpected error ({exc}) — keeping originals for this batch.")
    return None


# ─── OpenCode Zen backend ────────────────────────────────────────────────────

def _opencode_batch(client, model_name: str, texts: list) -> list:
    prompt = _build_prompt(texts)
    max_retries = 6
    for attempt in range(max_retries):
        try:
            response = client.chat.completions.create(
                model=model_name,
                messages=[{"role": "user", "content": prompt}],
            )
            raw = (response.choices[0].message.content or "").strip()
            return _parse_response(raw, len(texts), texts)
        except Exception as exc:
            delay = _handle_error(exc, attempt, max_retries)
            if delay is None:
                return texts
            time.sleep(delay)
    return texts


def build_opencode_client(api_key: str):
    try:
        from openai import OpenAI
    except ImportError:
        print("ERROR: openai not installed. Run: pip install openai")
        sys.exit(1)
    return OpenAI(api_key=api_key, base_url=OPENCODE_BASE_URL)


# ─── Google Gemini backend ───────────────────────────────────────────────────

def _gemini_batch(client, model_name: str, texts: list) -> list:
    prompt = _build_prompt(texts)
    max_retries = 6
    for attempt in range(max_retries):
        try:
            response = client.models.generate_content(model=model_name, contents=prompt)
            raw = response.text.strip()
            return _parse_response(raw, len(texts), texts)
        except Exception as exc:
            delay = _handle_error(exc, attempt, max_retries)
            if delay is None:
                return texts
            time.sleep(delay)
    return texts


def build_gemini_client(api_key: str):
    try:
        from google import genai
    except ImportError:
        print("ERROR: google-genai not installed. Run: pip install google-genai")
        sys.exit(1)
    return genai.Client(api_key=api_key)


# ─── Core workbook translation ───────────────────────────────────────────────

def translate_workbook(wb: openpyxl.Workbook, translate_fn, model_name: str, batch_size: int) -> None:
    pending = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if should_translate(cell):
                    pending.append((sheet_name, cell.row, cell.column, cell.value))

    total = len(pending)
    texts = [item[3] for item in pending]
    total_batches = (total + batch_size - 1) // batch_size

    print(f"\n{'─' * 60}")
    print(f"  Sheets       : {len(wb.sheetnames)}")
    print(f"  Text cells   : {total}")
    print(f"  Batches      : {total_batches}  (batch size = {batch_size})")
    print(f"  Model        : {model_name}")
    print(f"  Started      : {datetime.now().strftime('%H:%M:%S')}")
    print(f"{'─' * 60}\n")

    progress = Progress(total, total_batches)
    translated = []

    for i in range(0, total, batch_size):
        batch = texts[i : i + batch_size]
        batch_num = i // batch_size + 1
        sheet_name = pending[i][0]

        progress.start_batch(batch_num, sheet_name)
        result = translate_fn(batch)
        translated.extend(result)
        progress.finish_batch(len(batch))

        if i + batch_size < total:
            time.sleep(1.0)

    progress.done()

    for (sheet_name, row, col, _), new_value in zip(pending, translated):
        cell = wb[sheet_name].cell(row=row, column=col)
        if not isinstance(cell, MergedCell):
            cell.value = new_value

    print(f"\n{'─' * 60}")
    print(f"  Finished     : {datetime.now().strftime('%H:%M:%S')}")
    print(f"{'─' * 60}\n")


# ─── Entry point ────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Translate an Excel (.xlsx) file from Portuguese to English."
    )
    parser.add_argument("input",  help="Path to the source .xlsx file")
    parser.add_argument("output", help="Path for the translated .xlsx file (date and model appended automatically)")
    parser.add_argument(
        "--provider",
        choices=[PROVIDER_OPENCODE, PROVIDER_GEMINI],
        default=PROVIDER_OPENCODE,
        help=f"Translation provider (default: {PROVIDER_OPENCODE})",
    )
    parser.add_argument(
        "--api-key",
        default=None,
        help=(
            f"API key for the selected provider. "
            f"Falls back to OPENCODE_ZEN_API_KEY ({PROVIDER_OPENCODE}) "
            f"or GEMINI_API_KEY ({PROVIDER_GEMINI}) env vars."
        ),
    )
    parser.add_argument(
        "--model",
        default=None,
        help=(
            f"Model name override. Defaults: "
            f"{PROVIDER_OPENCODE}={OPENCODE_DEFAULT!r}, "
            f"{PROVIDER_GEMINI}={GEMINI_DEFAULT!r}"
        ),
    )
    parser.add_argument(
        "--batch-size",
        type=int,
        default=20,
        help="Number of texts per API call (default: 20)",
    )
    parser.add_argument(
        "--list-models",
        action="store_true",
        help="List available models for the selected provider and exit",
    )
    args = parser.parse_args()

    # Resolve API key and model defaults per provider
    if args.provider == PROVIDER_OPENCODE:
        api_key = args.api_key or os.environ.get("OPENCODE_ZEN_API_KEY")
        model   = args.model or OPENCODE_DEFAULT
        env_var = "OPENCODE_ZEN_API_KEY"
        client  = None  # built after key check
    else:
        api_key = args.api_key or os.environ.get("GEMINI_API_KEY")
        model   = args.model or GEMINI_DEFAULT
        env_var = "GEMINI_API_KEY"
        client  = None

    if not api_key:
        print(
            f"ERROR: API key not found for provider '{args.provider}'.\n"
            f"  Pass --api-key YOUR_KEY  or  set the {env_var} environment variable."
        )
        sys.exit(1)

    # Build client
    if args.provider == PROVIDER_OPENCODE:
        client = build_opencode_client(api_key)
        translate_fn = lambda texts: _opencode_batch(client, model, texts)

        if args.list_models:
            print(f"Available models ({PROVIDER_OPENCODE}):\n")
            for m in client.models.list():
                print(f"  {m.id}")
            return

    else:
        client = build_gemini_client(api_key)
        translate_fn = lambda texts: _gemini_batch(client, model, texts)

        if args.list_models:
            print(f"Available models ({PROVIDER_GEMINI}):\n")
            for m in client.models.list():
                if "generateContent" in (m.supported_actions or []):
                    print(f"  {m.name}")
            return

    print(f"Provider : {args.provider}")
    print(f"Loading  : {args.input}")
    wb = openpyxl.load_workbook(args.input)

    translate_workbook(wb, translate_fn, model, args.batch_size)

    base, ext = os.path.splitext(args.output)
    date_str   = datetime.now().strftime("%Y-%m-%d")
    model_slug = model.replace("/", "-")
    output_path = f"{base}_{date_str}_{model_slug}{ext}"

    print(f"Saving: {output_path}")
    wb.save(output_path)
    print("Done!")


if __name__ == "__main__":
    main()
