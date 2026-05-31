import pytest
import re
import tempfile
import os

import openpyxl

from translate_excel import _sanitize, _sanitize_sheet_name


class TestSanitizeSheetName:
    def test_truncates_to_31_chars(self):
        long_name = "Undergraduate and graduate courses"
        assert len(long_name) == 34
        result = _sanitize_sheet_name(long_name)
        assert len(result) <= 31
        assert result == "Undergraduate and graduate cour"

    def test_strips_xml_illegal_chars(self):
        result = _sanitize_sheet_name("Sheet\x00Name\x01")
        assert "\x00" not in result
        assert "\x01" not in result

    def test_strips_excel_illegal_chars(self):
        cases = [
            ("A/B", "AB"),
            ("C?D", "CD"),
            ("E*F", "EF"),
            ("G[H]", "GH"),
            ("I:J", "IJ"),
            ("K\\L", "KL"),
        ]
        for inp, expected in cases:
            assert _sanitize_sheet_name(inp) == expected

    def test_strips_mixed_illegal_chars(self):
        result = _sanitize_sheet_name("A/B?C:D")
        assert result == "ABCD"

    def test_strips_leading_trailing_quotes(self):
        assert _sanitize_sheet_name("'My Sheet'") == "My Sheet"

    def test_fallback_to_sheet_when_empty(self):
        assert _sanitize_sheet_name("\x00\x01") == "Sheet"

    def test_fallback_to_sheet_when_only_illegal(self):
        assert _sanitize_sheet_name("[]:*?") == "Sheet"

    def test_preserves_valid_names(self):
        result = _sanitize_sheet_name("Valid Sheet Name")
        assert result == "Valid Sheet Name"

    def test_normal_name_within_limit(self):
        name = "Continuing Education"
        assert _sanitize_sheet_name(name) == name

    def test_edge_case_exactly_31_chars(self):
        name = "A" * 31
        result = _sanitize_sheet_name(name)
        assert len(result) == 31
        assert result == name

    def test_edge_case_31_chars_with_illegal_stripped(self):
        name = "A" * 31 + ":"
        result = _sanitize_sheet_name(name)
        assert len(result) == 31
        assert ":" not in result


class TestSanitizeCellValue:
    def test_removes_xml_illegal_chars(self):
        assert _sanitize("Hello\x00World") == "HelloWorld"
        assert _sanitize("\x01\x02\x03") == ""

    def test_preserves_normal_text(self):
        assert _sanitize("Hello World") == "Hello World"

    def test_preserves_newlines_and_tabs(self):
        assert _sanitize("Line1\nLine2\tTab") == "Line1\nLine2\tTab"

    def test_preserves_utf8(self):
        assert _sanitize("Café résumé") == "Café résumé"

    def test_preserves_xml_special_chars(self):
        assert _sanitize("A & B < C > D \"E\" F'") == "A & B < C > D \"E\" F'"


class TestSheetRenamingInWorkbook:
    def create_workbook_with_names(self, names):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for name in names:
            ws = wb.create_sheet(title=name)
        return wb

    def test_sheet_name_truncated_on_save(self):
        wb = self.create_workbook_with_names(["Disciplinas de graduação e pós"])
        ws = wb.active
        long_new = "Undergraduate and graduate courses"
        ws.title = _sanitize_sheet_name(long_new)
        assert len(ws.title) <= 31
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        try:
            wb.save(tmp.name)
            wb.close()

            wb2 = openpyxl.load_workbook(tmp.name)
            assert wb2.sheetnames[0] == "Undergraduate and graduate cour"
            wb2.close()
        finally:
            os.unlink(tmp.name)

    def _apply_dedup(self, names: list[str]) -> list[str]:
        used = set()
        result = []
        for n in names:
            n = _sanitize_sheet_name(n)
            original = n
            suffix = 2
            while n in used:
                suffix_str = f" ({suffix})"
                max_base = 31 - len(suffix_str)
                n = f"{original[:max_base]}{suffix_str}"
                suffix += 1
            used.add(n)
            result.append(n)
        return result

    def test_duplicate_sheet_names_deduped(self):
        result = self._apply_dedup(["Data", "Data"])
        assert result == ["Data", "Data (2)"]

    def test_multiple_duplicate_sheet_names_deduped(self):
        result = self._apply_dedup(["Data", "Data", "Data"])
        assert result == ["Data", "Data (2)", "Data (3)"]

    def test_realistic_workbook_save_and_reload(self):
        orig_names = [
            "Educação Continuada",
            "Disciplinas de graduação e pós",
        ]
        translated = [
            "Continuing Education",
            "Undergraduate and graduate courses",
        ]
        wb = self.create_workbook_with_names(orig_names)
        used = set()
        for ws, new in zip(wb.worksheets, translated):
            new = _sanitize_sheet_name(new)
            original_name = new
            suffix = 2
            while new in used:
                suffix_str = f" ({suffix})"
                max_base = 31 - len(suffix_str)
                new = f"{original_name[:max_base]}{suffix_str}"
                suffix += 1
            used.add(new)
            ws.title = new

        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        try:
            wb.save(tmp.name)
            wb.close()

            wb2 = openpyxl.load_workbook(tmp.name)
            assert wb2.sheetnames[0] == "Continuing Education"
            assert wb2.sheetnames[1] == "Undergraduate and graduate cour"
            wb2.close()
        finally:
            os.unlink(tmp.name)


class TestTableColumnDedup:
    def test_duplicate_names_get_suffix(self):
        names = ["Name", "Name"]
        seen = set()
        result = []
        for n in names:
            original = n
            suffix = 2
            while n in seen:
                n = f"{original}_{suffix}"
                suffix += 1
            seen.add(n)
            result.append(n)
        assert result == ["Name", "Name_2"]

    def test_triplicate_names(self):
        names = ["Code", "Code", "Code"]
        seen = set()
        result = []
        for n in names:
            original = n
            suffix = 2
            while n in seen:
                n = f"{original}_{suffix}"
                suffix += 1
            seen.add(n)
            result.append(n)
        assert result == ["Code", "Code_2", "Code_3"]

    def test_no_duplicates_unchanged(self):
        names = ["Name", "Code", "Type"]
        seen = set()
        result = []
        for n in names:
            original = n
            suffix = 2
            while n in seen:
                n = f"{original}_{suffix}"
                suffix += 1
            seen.add(n)
            result.append(n)
        assert result == ["Name", "Code", "Type"]
